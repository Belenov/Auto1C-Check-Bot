import re
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import time
import datetime
from contextlib import contextmanager
import pandas as pd
import warnings
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from threading import Thread, Lock
import os
import sys
import random
import json
import imaplib
import email
from email.header import decode_header
from telegram import Update, Bot, InlineKeyboardButton, InlineKeyboardMarkup, InputFile
from telegram.ext import (
    Updater,
    CommandHandler,
    MessageHandler,
    filters,
    CallbackContext,
    ConversationHandler,
    CallbackQueryHandler
)



MAX_TEXT_LENGTH = 100

LOGIN_URL = 'https://login.1c.ru/login'
DATA_URL = 'https://releases.1c.ru'

EXCEL_FILE_PATH = "C:/otchet.xlsx"
SHEET_NAME = "Актуальные версии конфигураций"
REPORT_FOLDER = "C:/"

USERNAME = ''
PASSWORD = ''

EMAIL_HOST = "imap.mail.ru"   
EMAIL_USER = ""   
EMAIL_PASS = ""           
EMAIL_FOLDER = "INBOX"              

excel_lock = Lock()

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

BOT_TOKEN = ""
TELEGRAM_AUTH_PASSWORD = ""

AUTHENTICATED_USERS = set()
SUBSCRIBED_USERS = set()

UPDATE_INTERVAL = 86400

SUBSCRIBERS_FILE = "C:/subscribers.json"
telegram_bot = None


def load_subscribers():
    if os.path.exists(SUBSCRIBERS_FILE):
        try:
            with open(SUBSCRIBERS_FILE, "r") as f:
                data = json.load(f)
                return set(data)
        except Exception as e:
            print("Ошибка при загрузке подписчиков:", e)
    return set()

def save_subscribers():
    try:
        with open(SUBSCRIBERS_FILE, "w") as f:
            json.dump(list(SUBSCRIBED_USERS), f)
    except Exception as e:
        print("Ошибка при сохранении подписчиков:", e)

SUBSCRIBED_USERS = load_subscribers()

def load_releases_from_excel(excel_file=EXCEL_FILE_PATH, sheet_name=SHEET_NAME):
    with excel_lock:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook[sheet_name]
        releases = {}
        for row in range(3, sheet.max_row + 1):
            if row == 7:
                continue
            config_name = sheet.cell(row=row, column=1).value
            current_version = sheet.cell(row=row, column=3).value
            if config_name and current_version:
                releases[config_name.strip()] = {'row': row, 'version': current_version.strip()}
        return releases, workbook, sheet

def update_releases_in_excel(releases, workbook, sheet_name=SHEET_NAME):
    with excel_lock:
        sheet = workbook[sheet_name]
        for name, data in releases.items():
            sheet.cell(row=data['row'], column=3, value=data['version'])
        workbook.save(EXCEL_FILE_PATH)
        print(f"Файл {EXCEL_FILE_PATH} успешно обновлен.")


def extract_first_version_from_html(version_column):
    version_links = version_column.find_all('a')
    return version_links[0].get_text(strip=True) if version_links else None

def truncate_text(text, max_length=MAX_TEXT_LENGTH):
    return text[:max_length] + "..." if text and len(text) > max_length else text

def filter_version(version):
    return re.sub(r'[^0-9.]', '', version)

def extract_first_version(version_string):
    return version_string.split(' ', 1)[0].strip()

def compare_versions(current_version, new_version):

    current_version = filter_version(current_version)
    new_version = filter_version(new_version)
    try:
        current_parts = list(map(int, current_version.split('.')))
        new_parts = list(map(int, new_version.split('.')))
        max_length = max(len(current_parts), len(new_parts))
        current_parts.extend([0] * (max_length - len(current_parts)))
        new_parts.extend([0] * (max_length - len(current_parts)))
        return new_parts > current_parts
    except ValueError:
        print(f"Ошибка сравнения версий: '{current_version}' vs '{new_version}'")
        return False


@contextmanager
def selenium_driver():
    service = Service(ChromeDriverManager().install())
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    driver = webdriver.Chrome(service=service, options=options)
    try:
        yield driver
    finally:
        driver.quit()

def login(driver):
    driver.get(LOGIN_URL)
    driver.find_element(By.ID, "username").send_keys(USERNAME)
    driver.find_element(By.ID, "password").send_keys(PASSWORD + Keys.RETURN)
    if driver.current_url == LOGIN_URL:
        raise Exception("Ошибка авторизации!")
    print("Авторизация успешна!")

def check_updates():
    with selenium_driver() as driver:
        login(driver)
        driver.get(DATA_URL)
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        table_body = soup.find('tbody')
        if not table_body:
            print("Данные не найдены.")
            return []
        releases_dict, workbook, sheet = load_releases_from_excel()
        updated_products = []

        for row in table_body.find_all('tr'):
            cells = row.find_all('td')
            if len(cells) >= 3:
                name = cells[0].get_text(strip=True)
                current_version = extract_first_version_from_html(cells[1])
                if current_version and name in releases_dict:
                    old_version = releases_dict[name]['version']
                    if compare_versions(old_version, current_version):
                        releases_dict[name]['version'] = current_version
                        updated_products.append(f"Обновлено: {name} - {current_version}")

        if updated_products:
            update_releases_in_excel(releases_dict, workbook)
            print("\n".join(updated_products))
        else:
            print("Обновлений нет.")
        return updated_products

def check_email_updates():

    print("Проверяем новые письма...")

    try:
        mail = imaplib.IMAP4_SSL(EMAIL_HOST)
        mail.login(EMAIL_USER, EMAIL_PASS)
        mail.select(EMAIL_FOLDER)
    except Exception as e:
        print(f"Не удалось подключиться к почтовому серверу: {e}")
        return []

    criterion = '(ALL)'

    try:
        result, data = mail.search(None, criterion)
        if result != 'OK':
            print("Не удалось выполнить поиск писем.")
            mail.logout()
            return []

        email_ids = data[0].split()
        parsed_results = []

        for eid in email_ids:
            res, msg_data = mail.fetch(eid, '(RFC822)')
            if res != 'OK':
                continue

            raw_email = msg_data[0][1]
            try:
                msg = email.message_from_bytes(raw_email)
            except:
                continue

            subject, encoding = decode_header(msg["Subject"])[0]
            if isinstance(subject, bytes):
                subject = subject.decode(encoding if encoding else 'utf-8', errors='ignore')

            body_text = ""
            if msg.is_multipart():
                for part in msg.walk():
                    ctype = part.get_content_type()
                    cdispo = str(part.get('Content-Disposition'))
                    if ctype == 'text/plain' and 'attachment' not in cdispo:
                        charset = part.get_content_charset()
                        try:
                            body_text = part.get_payload(decode=True).decode(charset if charset else 'utf-8', errors='ignore')
                        except:
                            pass
            else:
                body_text = msg.get_payload(decode=True)
                if body_text:
                    try:
                        body_text = body_text.decode('utf-8', errors='ignore')
                    except:
                        body_text = str(body_text)

            updates_count = None
            match = re.search(r'(\d+)\s+шт', subject, re.IGNORECASE)
            if match:
                updates_count = int(match.group(1))
            warnings_found = False
            errors_found = False

            if "предупреждение" in body_text.lower():
                warnings_found = True
            if "ошибка" in body_text.lower():
                errors_found = True

            parsed_results.append({
                "subject": subject,
                "updates_count": updates_count,
                "warnings_found": warnings_found,
                "errors_found": errors_found,
                "body_snippet": body_text[:200]
            })

        mail.close()
        mail.logout()
    except Exception as e:
        print(f"Ошибка при чтении писем: {e}")
        return []

    if parsed_results:
        save_email_updates_to_excel(parsed_results)

    for item in parsed_results:
        if item["warnings_found"] or item["errors_found"]:
            message = (
                f"Внимание!\n"
                f"В письме «{item['subject']}» обнаружено предупреждение или ошибка.\n"
                f"Кол-во обновлений (если указано): {item['updates_count']}"
            )
            for chat_id in SUBSCRIBED_USERS:
                try:
                    telegram_bot.send_message(chat_id=chat_id, text=message)
                except Exception as e:
                    print(f"Ошибка отправки уведомления пользователю {chat_id}: {e}")

    return parsed_results

def save_email_updates_to_excel(parsed_results):
    with excel_lock:
        if not os.path.exists(EXCEL_FILE_PATH):
            wb = Workbook()
            wb.save(EXCEL_FILE_PATH)

        wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
        sheet_name = "MailUpdates"

        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
            ws = wb[sheet_name]
            ws.append(["Дата/Время", "Тема письма", "Кол-во обновлений", "Предупреждение", "Ошибка", "Текст (фрагмент)"])
        else:
            ws = wb[sheet_name]

        for item in parsed_results:
            now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([
                now_str,
                item["subject"],
                item["updates_count"] if item["updates_count"] is not None else "",
                "Да" if item["warnings_found"] else "",
                "Да" if item["errors_found"] else "",
                item["body_snippet"]
            ])

        wb.save(EXCEL_FILE_PATH)
    print(f"Записано {len(parsed_results)} писем в лист '{sheet_name}' файла {EXCEL_FILE_PATH}.")


def snake_game():
    if os.name != 'nt':
        print("Игра доступна только на Windows.")
        input("Нажмите Enter для возврата в меню...")
        return

    import msvcrt
    from colorama import init, Fore, Back, Style
    init()

    width = 21
    height = 40
    num_walls = 30
    colors = {
        'head': Fore.GREEN + Style.BRIGHT,
        'body': Fore.GREEN,
        'food': Fore.RED + Style.BRIGHT,
        'wall': Fore.YELLOW + Style.BRIGHT,
        'tile': Fore.WHITE + Style.DIM,
        'reset': Style.RESET_ALL
    }
    tiles = ['░', '▒', '▓']

    snake = [(width//2, height//2)]
    direction = (1, 0)
    score = 0
    walls = set()

    def generate_walls():
        walls = set()
        while len(walls) < num_walls:
            x = random.randint(0, width-1)
            y = random.randint(0, height-1)
            if (x, y) not in snake and (x, y) != snake[0]:
                walls.add((x, y))
        return walls

    def generate_food():
        while True:
            pos = (random.randint(0, width-1), random.randint(0, height-1))
            if pos not in walls and pos not in snake:
                return pos
    walls = generate_walls()
    food = generate_food()

    while True:
        os.system('cls')
        for y in range(height):
            line = []
            for x in range(width):
                if (x, y) in walls:
                    line.append(f"{colors['wall']}#{colors['reset']}")
                elif (x, y) == snake[0]:
                    line.append(f"{colors['head']}●{colors['reset']}")
                elif (x, y) in snake:
                    line.append(f"{colors['body']}○{colors['reset']}")
                elif (x, y) == food:
                    line.append(f"{colors['food']}◆{colors['reset']}")
                else:
                    tile = random.choice(tiles)
                    line.append(f"{colors['tile']}{tile}{colors['reset']}")
            print(''.join(line))

        print(f"{Style.BRIGHT}Счет: {score}{Style.RESET_ALL}")
        print(f"Управление: WASD | Q - выход")

        key = None
        if msvcrt.kbhit():
            key = msvcrt.getch().decode('utf-8').lower()

        new_dir = direction
        if key == 'w' and direction != (0, 1):
            new_dir = (0, -1)
        elif key == 's' and direction != (0, -1):
            new_dir = (0, 1)
        elif key == 'a' and direction != (1, 0):
            new_dir = (-1, 0)
        elif key == 'd' and direction != (-1, 0):
            new_dir = (1, 0)
        elif key == 'q':
            break

        direction = new_dir
        new_head = (snake[0][0] + direction[0], snake[0][1] + direction[1])

        if (
            new_head[0] < 0 
            or new_head[0] >= width 
            or new_head[1] < 0 
            or new_head[1] >= height 
            or new_head in walls 
            or new_head in snake
        ):
            print(f"{Fore.RED}Game Over!{Style.RESET_ALL}")
            break
        snake.insert(0, new_head)
        if new_head == food:
            score += 10
            food = generate_food()
        else:
            snake.pop()

        time.sleep(0.15)

    input("Нажмите Enter для возврата в меню...")

def load_data():
    with excel_lock:
        df_releases = pd.read_excel(EXCEL_FILE_PATH, sheet_name='Актуальные версии конфигураций')
        df_updates = pd.read_excel(EXCEL_FILE_PATH, sheet_name='Список обновлений')
    return df_releases, df_updates

def filter_configurations(df_releases, choice):
    config_map = {
        '1': 'БУ (Бюджет)',
        '2': 'ХО (Хозрасчет)'
    }
    if choice in config_map:
        return df_releases[df_releases['Вид расчета'] == config_map[choice]]
    return df_releases

def process_report(choice=None):
    df_releases, df_updates = load_data()

    if choice is None:
        print("\nВыберите тип конфигурации:")
        print("1 - Бюджетные учреждения")
        print("2 - Хозрасчетные организации")
        print("3 - Все конфигурации")
        choice = input("Ваш выбор (1/2/3): ").strip()

    df_releases = filter_configurations(df_releases, choice)
    releases_dict = dict(zip(df_releases['Конфигурации'], df_releases['Версия']))

    def check_version(row):
        product = row['Программный продукт']
        version = row['Новый']
        if product in releases_dict:
            return 'Да' if version == releases_dict[product] else 'Нет'
        return None

    df_updates['Актуальный'] = df_updates.apply(check_version, axis=1)
    df_report = df_updates.dropna(subset=['Актуальный'])[['Клиент', 'Программный продукт', 'Новый', 'Актуальный']]

    summary_data = []
    exclude = ["Конфигурации хозрасчетных организаций", "Платформа 1С", "Конфигурации бюджетных учреждений"]

    for config in df_releases['Конфигурации'].unique():
        if config not in exclude:
            current_ver = releases_dict.get(config, 'Неизвестно')
            count = df_report[(df_report['Программный продукт'] == config) & (df_report['Актуальный'] == 'Да')].shape[0]
            summary_data.append([config, current_ver, count])

    summary_df = pd.DataFrame(summary_data, columns=['Конфигурация', 'Актуальная версия', 'Обновлено'])
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    report_path = f"{REPORT_FOLDER}report_{timestamp}.xlsx"

    with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
        df_report.to_excel(writer, sheet_name='Отчет', index=False)
        sheet = writer.sheets['Отчет']

        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        for row in sheet.iter_rows(min_row=2, max_row=len(df_report)+1):
            status = row[3].value
            fill_color = green_fill if status == 'Да' else red_fill if status == 'Нет' else None
            if fill_color:
                for cell in row:
                    cell.fill = fill_color

        summary_start = len(df_report) + 3
        summary_df.to_excel(writer, sheet_name='Отчет', startrow=summary_start, index=False)

    print(f"\nОтчет сохранен: {report_path}")
    print(f"Всего записей: {len(df_report)}")
    print(f"Актуальные: {df_report['Актуальный'].value_counts().get('Да', 0)}")
    print(f"Устаревшие: {df_report['Актуальный'].value_counts().get('Нет', 0)}")
    return report_path

def check_updates_loop():
    while True:
        print(f"\n[{datetime.datetime.now()}] Запуск проверки обновлений...")
        try:
            updates = check_updates()
            email_updates = check_email_updates()

            if updates and telegram_bot:
                message = "Обновления по 1С:\n" + "\n".join(updates)
                for chat_id in SUBSCRIBED_USERS:
                    try:
                        telegram_bot.send_message(chat_id=chat_id, text=message)
                    except Exception as e:
                        print(f"Ошибка отправки уведомления пользователю {chat_id}: {e}")

        except Exception as e:
            print(f"Ошибка при проверке обновлений: {str(e)}")

        print(f"Ожидаем {UPDATE_INTERVAL} секунд до следующей проверки...\n")
        time.sleep(UPDATE_INTERVAL)

def start(update: Update, context: CallbackContext):
    update.message.reply_text(
        "Привет! Добро пожаловать в бот обновлений.\n"
        "Для доступа используйте /login\n"
        "Доступные команды:\n"
        "/login - авторизация\n"
        "/logout - выйти из системы\n"
        "/report - создать отчёт\n"
        "/subscribe - подписаться на уведомления о новых версиях\n"
        "/unsubscribe - отписаться от уведомлений\n"
        "/check - ручная проверка обновлений\n"
        "/setinterval - изменить интервал проверки обновлений\n"
        "/help - список команд"
    )

LOGIN_STATE = 1

def login_command(update: Update, context: CallbackContext):
    update.message.reply_text("Пожалуйста, введите пароль для авторизации:")
    return LOGIN_STATE

def login_receive_password(update: Update, context: CallbackContext):
    password = update.message.text.strip()
    if password == TELEGRAM_AUTH_PASSWORD:
        AUTHENTICATED_USERS.add(update.effective_chat.id)
        update.message.reply_text("Аутентификация прошла успешно!")
        return ConversationHandler.END
    else:
        update.message.reply_text("Неверный пароль. Попробуйте ещё раз или введите /cancel для отмены.")
        return LOGIN_STATE

def cancel(update: Update, context: CallbackContext):
    update.message.reply_text("Авторизация отменена.")
    return ConversationHandler.END

def logout_command(update: Update, context: CallbackContext):
    if update.effective_chat.id in AUTHENTICATED_USERS:
        AUTHENTICATED_USERS.remove(update.effective_chat.id)
        update.message.reply_text("Вы вышли из системы.")
    else:
        update.message.reply_text("Вы не авторизованы.")

def report_command(update: Update, context: CallbackContext):
    if update.effective_chat.id not in AUTHENTICATED_USERS:
        update.message.reply_text("Сначала авторизуйтесь командой /login.")
        return
    keyboard = [
        [InlineKeyboardButton("Бюджетные учреждения", callback_data="report_choice:1")],
        [InlineKeyboardButton("Хозрасчетные организации", callback_data="report_choice:2")],
        [InlineKeyboardButton("Все конфигурации", callback_data="report_choice:3")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    update.message.reply_text("Выберите тип конфигурации:", reply_markup=reply_markup)

def report_callback(update: Update, context: CallbackContext):
    query = update.callback_query
    query.answer()
    data = query.data
    choice = data.split(":")[1]
    query.edit_message_text(text=f"Генерируется отчёт для выбранного типа конфигурации...")
    try:
        report_path = process_report(choice)
        with open(report_path, 'rb') as f:
            context.bot.send_document(chat_id=query.message.chat_id, document=f, filename=os.path.basename(report_path))
        context.bot.send_message(chat_id=query.message.chat_id, text="Отчёт отправлен.")
    except Exception as e:
        context.bot.send_message(chat_id=query.message.chat_id, text=f"Ошибка при создании отчёта: {str(e)}")

def subscribe_command(update: Update, context: CallbackContext):
    if update.effective_chat.id not in AUTHENTICATED_USERS:
        update.message.reply_text("Сначала авторизуйтесь командой /login.")
        return
    SUBSCRIBED_USERS.add(update.effective_chat.id)
    save_subscribers()
    update.message.reply_text("Вы подписаны на уведомления о новых версиях (включая почтовые).")

def unsubscribe_command(update: Update, context: CallbackContext):
    if update.effective_chat.id in SUBSCRIBED_USERS:
        SUBSCRIBED_USERS.remove(update.effective_chat.id)
        save_subscribers()
        update.message.reply_text("Вы отписаны от уведомлений.")
    else:
        update.message.reply_text("Вы не подписаны на уведомления.")

def help_command(update: Update, context: CallbackContext):
    update.message.reply_text(
        "Доступные команды:\n"
        "/login - авторизация\n"
        "/logout - выйти из системы\n"
        "/report - создать отчёт\n"
        "/subscribe - подписаться на уведомления о новых версиях\n"
        "/unsubscribe - отписаться от уведомлений\n"
        "/check - ручная проверка обновлений\n"
        "/setinterval - изменить интервал проверки обновлений\n"
        "/help - помощь"
    )

def manual_update_command(update: Update, context: CallbackContext):
    if update.effective_chat.id not in AUTHENTICATED_USERS:
        update.message.reply_text("Сначала авторизуйтесь командой /login.")
        return

    update.message.reply_text("Запуск ручной проверки обновлений...")

    try:
        updates = check_updates()
        email_updates = check_email_updates()

        msg_parts = []
        if updates:
            msg_parts.append("Обновления по 1С:\n" + "\n".join(updates))
        else:
            msg_parts.append("Обновлений 1С нет.")

        if email_updates:
            msg_parts.append(f"Обработано новых писем: {len(email_updates)}. Смотрите лист 'MailUpdates'.")
        else:
            msg_parts.append("Новых писем (подходящих под критерии) нет.")

        update.message.reply_text("\n\n".join(msg_parts))

    except Exception as e:
        update.message.reply_text(f"Ошибка при проверке обновлений: {str(e)}")

def set_interval_command(update: Update, context: CallbackContext):
    if update.effective_chat.id not in AUTHENTICATED_USERS:
        update.message.reply_text("Сначала авторизуйтесь командой /login.")
        return
    args = context.args
    if not args:
        update.message.reply_text("Использование: /setinterval <секунд>")
        return
    try:
        global UPDATE_INTERVAL
        new_interval = int(args[0])
        UPDATE_INTERVAL = new_interval
        update.message.reply_text(f"Интервал проверки обновлений изменен на {new_interval} секунд.")
    except ValueError:
        update.message.reply_text("Пожалуйста, введите корректное число секунд.")

def main():
    update_thread = Thread(target=check_updates_loop, daemon=True)
    update_thread.start()
    updater = Updater(BOT_TOKEN)
    dispatcher = updater.dispatcher
    dispatcher.add_handler(CommandHandler("start", start))

    login_conv_handler = ConversationHandler(
        entry_points=[CommandHandler("login", login_command)],
        states={
            LOGIN_STATE: [MessageHandler(filters.Filters.text & ~filters.Filters.command, login_receive_password)]
        },
        fallbacks=[CommandHandler("cancel", cancel)]
    )
    
    dispatcher.add_handler(login_conv_handler)
    dispatcher.add_handler(CommandHandler("logout", logout_command))
    dispatcher.add_handler(CommandHandler("report", report_command))
    dispatcher.add_handler(CallbackQueryHandler(report_callback, pattern=r'^report_choice:'))
    dispatcher.add_handler(CommandHandler("subscribe", subscribe_command))
    dispatcher.add_handler(CommandHandler("unsubscribe", unsubscribe_command))
    dispatcher.add_handler(CommandHandler("help", help_command))
    dispatcher.add_handler(CommandHandler("check", manual_update_command))
    dispatcher.add_handler(CommandHandler("setinterval", set_interval_command))

    updater.start_polling()

    global telegram_bot
    telegram_bot = updater.bot
    print("Telegram-бот запущен.")

    while True:
        command = input("\nВведите команду:\n1 - Создать отчёт\n2 - Выход\n3 - Играть в змейку\n> ").strip()
        if command == '1':
            process_report()
        elif command == '2':
            print("Завершение работы...")
            updater.stop()
            break
        elif command == '3':
            snake_game()
        else:
            print("Некорректный ввод")

if __name__ == "__main__":
    main()
